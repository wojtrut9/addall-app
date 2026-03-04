"""
excel_export.py — Add All Excel Report Generator
Generuje 2 pliki Excel:
  1. Pełna analiza (6 arkuszy)
  2. Prosta lista zamówień (do wgrania do iBiznes)
"""
import io
from datetime import datetime

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


# ── Kolory arkuszy ────────────────────────────────────────────────────────────
SHEET_COLORS = {
    "🚨 ZAMÓW DZIŚ":    "C0392B",   # czerwony
    "🟡 Zamów tydzień": "D4AC0D",   # żółty
    "🔵 W drodze":      "2980B9",   # niebieski
    "📈 Top movers":    "27AE60",   # zielony
    "⚫ Dead stock":    "555555",   # szary
    "📊 Pełna analiza": "2C3E50",   # granatowy
}

ROW_COLORS = {
    "ZAMÓW DZIŚ":   "FFE0E0",
    "ZAMÓW TYDZIEŃ": "FFFDE7",
    "DEAD STOCK":   "EEEEEE",
    "OK":           "FFFFFF",
}

RENAME_MAP = {
    "srednie_dzienne":      "Zuż/dzień",
    "dni_do_wyczerpania":   "Starczy (dni)",
    "ile_zamowic":          "Zamów (szt)",
    "wartosc_zamowienia":   "Wartość PLN",
    "wartosc_stanu":        "Wartość stanu PLN",
    "marza_pct":            "Marża %",
    "ilosc_sprzedana":      "Sprzedano (szt)",
    "wartosc_sprzedana":    "Sprzedano (PLN)",
    "liczba_transakcji":    "Transakcji WZ",
    "ostatnia_sprzedaz":    "Ostatnia sprzedaż",
    "prognoza_30d":         "Prognoza 30d",
    "safety_7d":            "Safety stock (7d)",
}


def _find_col(df, *hints):
    for hint in hints:
        m = next((c for c in df.columns if hint.lower() in c.lower()), None)
        if m:
            return m
    return None


def _style_sheet(ws, header_hex="366092"):
    """Formatuje nagłówek, szerokości kolumn i mrozi wiersz 1."""
    header_fill = PatternFill(start_color=header_hex, end_color=header_hex, fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=10)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center

    ws.row_dimensions[1].height = 28

    # Automatyczna szerokość kolumn
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max(max_len + 3, 10), 45)

    ws.freeze_panes = "A2"

    # Naprzemienne kolory wierszy
    light = PatternFill(start_color="F7F7F7", end_color="F7F7F7", fill_type="solid")
    for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
        if i % 2 == 0:
            for cell in row:
                if cell.fill.fill_type in (None, "none"):
                    cell.fill = light


def _get_base_cols(analiza):
    """Zwraca listę kolumn do wyświetlenia w tabelach."""
    candidates = [
        _find_col(analiza, "kod towaru / usługi", "kod towaru"),
        _find_col(analiza, "nazwa towaru"),
        _find_col(analiza, "dostawca"),
        "Stan",
        "Stan Min.",
        "srednie_dzienne",
        "dni_do_wyczerpania",
        "ile_zamowic",
        "wartosc_zamowienia",
        "marza_pct",
    ]
    return [c for c in candidates if c and c in analiza.columns]


def generate_full_excel(analiza, zam_df, summary):
    """
    Generuje pełny raport Excel z 6 arkuszami.
    Returns: bytes (zawartość pliku .xlsx)
    """
    output = io.BytesIO()
    base_cols = _get_base_cols(analiza)

    kod_col    = _find_col(analiza, "kod towaru / usługi", "kod towaru")
    nazwa_col  = _find_col(analiza, "nazwa towaru")
    dos_col    = _find_col(analiza, "dostawca")
    grupa_col  = _find_col(analiza, "grupa")

    with pd.ExcelWriter(output, engine="openpyxl") as writer:

        # ── Arkusz 1: ZAMÓW DZIŚ ─────────────────────────────────────────────
        dzis = (
            analiza[analiza["status"] == "ZAMÓW DZIŚ"]
            .sort_values("dni_do_wyczerpania")
        )
        _write_sheet(writer, dzis, base_cols, "🚨 ZAMÓW DZIŚ")

        # ── Arkusz 2: Zamów tydzień ──────────────────────────────────────────
        tydzien = (
            analiza[analiza["status"] == "ZAMÓW TYDZIEŃ"]
            .sort_values("dni_do_wyczerpania")
        )
        _write_sheet(writer, tydzien, base_cols, "🟡 Zamów tydzień")

        # ── Arkusz 3: W drodze ───────────────────────────────────────────────
        if zam_df is not None and len(zam_df) > 0:
            zam_clean = zam_df.drop(columns=["_data_realiz"], errors="ignore")
            zam_clean.to_excel(writer, sheet_name="🔵 W drodze", index=False)
        else:
            pd.DataFrame(
                {"Info": ["Brak otwartych zamówień lub nie wgrano pliku"]}
            ).to_excel(writer, sheet_name="🔵 W drodze", index=False)

        # ── Arkusz 4: Top movers ─────────────────────────────────────────────
        top_cols = [c for c in [
            kod_col, nazwa_col, dos_col,
            "srednie_dzienne", "Stan", "dni_do_wyczerpania",
            "marza_pct", "wartosc_stanu",
        ] if c and c in analiza.columns]
        top = (
            analiza[analiza["srednie_dzienne"] > 0]
            .nlargest(20, "srednie_dzienne")
        )
        _write_sheet(writer, top, top_cols, "📈 Top movers")

        # ── Arkusz 5: Dead stock ─────────────────────────────────────────────
        dead_cols = [c for c in [
            kod_col, nazwa_col, dos_col,
            "Stan", "wartosc_stanu", "ostatnia_sprzedaz",
        ] if c and c in analiza.columns]
        dead = (
            analiza[analiza["status"] == "DEAD STOCK"]
            .sort_values("wartosc_stanu", ascending=False)
        )
        _write_sheet(writer, dead, dead_cols, "⚫ Dead stock")

        # ── Arkusz 6: Pełna analiza ──────────────────────────────────────────
        full_cols = [c for c in [
            kod_col, nazwa_col, dos_col, grupa_col,
            "Stan", "Stan Min.", "srednie_dzienne", "dni_do_wyczerpania",
            "wartosc_stanu", "ile_zamowic", "wartosc_zamowienia",
            "ilosc_sprzedana", "wartosc_sprzedana", "marza_pct",
            "liczba_transakcji", "ostatnia_sprzedaz", "status",
        ] if c and c in analiza.columns]
        _write_sheet(writer, analiza, full_cols, "📊 Pełna analiza")

        # ── Stylowanie ───────────────────────────────────────────────────────
        wb = writer.book
        for sheet_name, color in SHEET_COLORS.items():
            if sheet_name in wb.sheetnames:
                _style_sheet(wb[sheet_name], color)

        # ── Arkusz podsumowania (na początku) ────────────────────────────────
        _add_summary_sheet(writer, summary)
        wb = writer.book
        # Przesuń summary na 1. miejsce
        if "📋 Podsumowanie" in wb.sheetnames:
            wb.move_sheet("📋 Podsumowanie", offset=-(len(wb.sheetnames) - 1))
        _style_sheet(wb["📋 Podsumowanie"], "2C3E50")

    output.seek(0)
    return output.read()


def _write_sheet(writer, df, cols, sheet_name):
    """Zapisuje DataFrame do arkusza z przemianowaniem kolumn."""
    available = [c for c in cols if c in df.columns]
    df[available].rename(columns=RENAME_MAP).to_excel(
        writer, sheet_name=sheet_name, index=False
    )


def _add_summary_sheet(writer, summary):
    """Dodaje arkusz podsumowania z kluczowymi metrykami."""
    data = {
        "Metryka": [
            "Data analizy",
            "Okres danych",
            "Produktów w bazie",
            "Produktów aktywnych (ruch ≥0.3 szt/dzień)",
            "Wartość magazynu (zakup netto)",
            "─────────────",
            "Produktów do zamówienia DZIŚ",
            "Wartość zamówień DZIŚ",
            "Produktów do zamówienia w tygodniu",
            "Wartość zamówień tydzień",
            "─────────────",
            "Dead stock — produktów",
            "Dead stock — zamrożony kapitał",
        ],
        "Wartość": [
            summary["data_analizy"],
            f"{summary['data_od']} — {summary['data_do']} ({summary['dni_okresu']} dni)",
            summary["produktow_total"],
            summary["produktow_aktywnych"],
            f"{summary['wartosc_magazynu']:,.0f} PLN".replace(",", " "),
            "",
            summary["produktow_dzis"],
            f"{summary['wartosc_dzis']:,.0f} PLN".replace(",", " "),
            summary["produktow_tydzien"],
            f"{summary['wartosc_tydzien']:,.0f} PLN".replace(",", " "),
            "",
            summary["dead_stock_produktow"],
            f"{summary['dead_stock_wartosc']:,.0f} PLN".replace(",", " "),
        ],
    }
    pd.DataFrame(data).to_excel(
        writer, sheet_name="📋 Podsumowanie", index=False
    )


def generate_order_excel(analiza):
    """
    Generuje prostą listę zamówień (plik do wgrania do iBiznes).
    Returns: bytes
    """
    output = io.BytesIO()

    kod_col   = _find_col(analiza, "kod towaru / usługi", "kod towaru")
    nazwa_col = _find_col(analiza, "nazwa towaru")
    dos_col   = _find_col(analiza, "dostawca")
    jm_col    = _find_col(analiza, "jm", "jednostka miary", "j.m.")

    to_order = analiza[
        analiza["status"].isin(["ZAMÓW DZIŚ", "ZAMÓW TYDZIEŃ"])
    ].copy()

    to_order["Priorytet"] = to_order["status"].map(
        {"ZAMÓW DZIŚ": "DZIŚ", "ZAMÓW TYDZIEŃ": "TYDZIEŃ"}
    )

    col_map = {}
    for src, dst in [
        (kod_col,   "Kod towaru"),
        (nazwa_col, "Nazwa"),
        ("ile_zamowic", "Ilość"),
        (jm_col, "JM"),
        ("Cena zakupu netto", "Cena zakupu netto"),
        (dos_col, "Dostawca"),
        ("Priorytet", "Priorytet"),
    ]:
        if src and src in to_order.columns:
            col_map[src] = dst

    result = to_order[list(col_map.keys())].rename(columns=col_map)

    sort_cols = [c for c in ["Priorytet", "Dostawca"] if c in result.columns]
    if sort_cols:
        result = result.sort_values(sort_cols)

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        result.to_excel(writer, sheet_name="Do zamówienia", index=False)

        wb = writer.book
        ws = wb["Do zamówienia"]
        _style_sheet(ws, "27AE60")

        # Koloruj wiersze: DZIŚ = czerwone, TYDZIEŃ = żółte
        red_fill    = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
        yellow_fill = PatternFill(start_color="FFFDE7", end_color="FFFDE7", fill_type="solid")

        prior_idx = None
        for i, cell in enumerate(ws[1], 1):
            if cell.value == "Priorytet":
                prior_idx = i
                break

        if prior_idx:
            for row in ws.iter_rows(min_row=2):
                val = row[prior_idx - 1].value
                fill = red_fill if val == "DZIŚ" else yellow_fill
                for cell in row:
                    cell.fill = fill

    output.seek(0)
    return output.read()
